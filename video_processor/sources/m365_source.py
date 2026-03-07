"""Microsoft 365 source connector using the m365 CLI (cli-microsoft365).

Fetches documents from SharePoint and OneDrive via the `m365` CLI tool.
Outputs plain text suitable for KG ingestion.

Requires: npm install -g @pnp/cli-microsoft365
Auth:     m365 login (interactive)
Docs:     https://pnp.github.io/cli-microsoft365/
"""

import json
import logging
import shutil
import subprocess
import tempfile
from pathlib import Path
from typing import Any, Dict, List, Optional

from video_processor.sources.base import BaseSource, SourceFile

logger = logging.getLogger(__name__)

# Document MIME types we can extract text from
_DOC_EXTENSIONS = {
    ".docx",
    ".doc",
    ".xlsx",
    ".xls",
    ".pptx",
    ".ppt",
    ".pdf",
    ".txt",
    ".md",
    ".csv",
    ".html",
    ".htm",
}


def _run_m365(args: List[str], timeout: int = 30) -> Any:
    """Run an m365 CLI command and return parsed JSON output."""
    cmd = ["m365"] + args + ["--output", "json"]
    proc = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)
    if proc.returncode != 0:
        raise RuntimeError(f"m365 {' '.join(args)} failed: {proc.stderr.strip()}")
    try:
        return json.loads(proc.stdout)
    except json.JSONDecodeError:
        return proc.stdout.strip()


class M365Source(BaseSource):
    """
    Fetch documents from SharePoint Online and OneDrive via the m365 CLI.

    Usage:
        # SharePoint site
        source = M365Source(
            web_url="https://contoso.sharepoint.com/sites/project-x",
            folder_url="/sites/project-x/Shared Documents"
        )

        # OneDrive
        source = M365Source(
            web_url="https://contoso-my.sharepoint.com/personal/user_contoso_com",
            folder_url="/personal/user_contoso_com/Documents"
        )

        files = source.list_videos()
        source.download_all(files, Path("./docs"))
    """

    def __init__(
        self,
        web_url: str,
        folder_url: Optional[str] = None,
        file_ids: Optional[List[str]] = None,
        recursive: bool = False,
    ):
        self.web_url = web_url
        self.folder_url = folder_url
        self.file_ids = file_ids or []
        self.recursive = recursive

    def authenticate(self) -> bool:
        """Check if m365 CLI is installed and logged in."""
        if not shutil.which("m365"):
            logger.error("m365 CLI not found. Install with: npm install -g @pnp/cli-microsoft365")
            return False
        try:
            result = _run_m365(["status"], timeout=10)
            # m365 status returns connection info when logged in
            if isinstance(result, dict) and result.get("connectedAs"):
                return True
            if isinstance(result, str) and "Logged in" in result:
                return True
            logger.error("m365 not logged in. Run: m365 login")
            return False
        except (RuntimeError, subprocess.TimeoutExpired):
            logger.error("m365 not logged in. Run: m365 login")
            return False

    def list_videos(
        self,
        folder_id: Optional[str] = None,
        folder_path: Optional[str] = None,
        patterns: Optional[List[str]] = None,
    ) -> List[SourceFile]:
        """List documents in SharePoint/OneDrive. Returns docs, not just videos."""
        files: List[SourceFile] = []

        # Fetch specific files by ID
        if self.file_ids:
            for fid in self.file_ids:
                try:
                    result = _run_m365(
                        [
                            "spo",
                            "file",
                            "get",
                            "--webUrl",
                            self.web_url,
                            "--id",
                            fid,
                        ]
                    )
                    files.append(_result_to_source_file(result))
                except RuntimeError as e:
                    logger.warning(f"Failed to get file {fid}: {e}")
            return files

        # List files in folder
        folder = folder_path or self.folder_url
        if not folder:
            logger.error("No folder URL specified. Use --folder-url or folder_path parameter.")
            return []

        try:
            args = [
                "file",
                "list",
                "--webUrl",
                self.web_url,
                "--folderUrl",
                folder,
            ]
            if self.recursive:
                args.append("--recursive")

            result = _run_m365(args, timeout=60)
        except RuntimeError as e:
            logger.error(f"Failed to list files: {e}")
            return []

        items = result if isinstance(result, list) else []
        for item in items:
            name = item.get("Name", item.get("name", ""))
            ext = Path(name).suffix.lower()
            if ext in _DOC_EXTENSIONS:
                files.append(_result_to_source_file(item))

        logger.info(f"Found {len(files)} document(s) in {folder}")
        return files

    def download(self, file: SourceFile, destination: Path) -> Path:
        """Download a file from SharePoint/OneDrive."""
        destination = Path(destination)
        destination.parent.mkdir(parents=True, exist_ok=True)

        args = [
            "spo",
            "file",
            "get",
            "--webUrl",
            self.web_url,
            "--asFile",
            "--path",
            str(destination),
        ]

        # Use URL if available in path field, otherwise use ID
        if file.path:
            args.extend(["--url", file.path])
        else:
            args.extend(["--id", file.id])

        _run_m365(args, timeout=120)
        logger.info(f"Downloaded {file.name} to {destination}")
        return destination

    def download_as_text(self, file: SourceFile) -> str:
        """Download a file and attempt to extract text content."""
        # For text-based formats, get as string directly
        text_exts = {".txt", ".md", ".csv", ".html", ".htm"}
        ext = Path(file.name).suffix.lower()

        if ext in text_exts:
            try:
                args = [
                    "spo",
                    "file",
                    "get",
                    "--webUrl",
                    self.web_url,
                    "--asString",
                ]
                if file.path:
                    args.extend(["--url", file.path])
                else:
                    args.extend(["--id", file.id])

                result = _run_m365(args, timeout=60)
                return result if isinstance(result, str) else json.dumps(result)
            except RuntimeError:
                pass

        # For binary formats, download to temp and extract
        with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
            tmp_path = Path(tmp.name)

        try:
            self.download(file, tmp_path)
            return _extract_text(tmp_path)
        finally:
            tmp_path.unlink(missing_ok=True)

    def fetch_all_text(self) -> Dict[str, str]:
        """List all docs and return {filename: text_content} dict."""
        files = self.list_videos()
        results = {}
        for f in files:
            try:
                results[f.name] = self.download_as_text(f)
            except Exception as e:
                logger.warning(f"Failed to fetch {f.name}: {e}")
                results[f.name] = f"[Error: {e}]"
        return results

    def collate(self, separator: str = "\n\n---\n\n") -> str:
        """Fetch all docs and collate into a single text blob for ingestion."""
        docs = self.fetch_all_text()
        parts = []
        for name, content in docs.items():
            parts.append(f"# {name}\n\n{content}")
        return separator.join(parts)


def _result_to_source_file(item: dict) -> SourceFile:
    """Convert an m365 file result to SourceFile."""
    name = item.get("Name", item.get("name", "Untitled"))
    file_id = item.get("UniqueId", item.get("uniqueId", item.get("id", "")))
    size = item.get("Length", item.get("length", item.get("size")))
    path = item.get("ServerRelativeUrl", item.get("serverRelativeUrl"))
    modified = item.get("TimeLastModified", item.get("lastModifiedDateTime"))

    return SourceFile(
        name=name,
        id=str(file_id),
        size_bytes=int(size) if size else None,
        mime_type=None,
        modified_at=modified,
        path=path,
    )


def _extract_text(path: Path) -> str:
    """Best-effort text extraction from a downloaded file."""
    ext = path.suffix.lower()

    if ext in {".txt", ".md", ".csv"}:
        return path.read_text(encoding="utf-8", errors="replace")

    if ext in {".html", ".htm"}:
        from video_processor.sources.web_source import _strip_html_tags

        return _strip_html_tags(path.read_text(encoding="utf-8", errors="replace"))

    if ext == ".pdf":
        try:
            import fitz  # pymupdf

            doc = fitz.open(str(path))
            return "\n\n".join(page.get_text() for page in doc)
        except ImportError:
            return f"[PDF file: {path.name} — install pymupdf to extract text]"

    if ext in {".docx", ".pptx", ".xlsx"}:
        # Try python-docx / openpyxl / python-pptx if available
        try:
            if ext == ".docx":
                from docx import Document

                doc = Document(str(path))
                return "\n\n".join(p.text for p in doc.paragraphs if p.text.strip())
            elif ext == ".xlsx":
                import openpyxl

                wb = openpyxl.load_workbook(str(path), read_only=True)
                rows = []
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    for row in ws.iter_rows(values_only=True):
                        cells = [str(c) if c is not None else "" for c in row]
                        if any(cells):
                            rows.append("\t".join(cells))
                return "\n".join(rows)
        except ImportError:
            return f"[{ext} file: {path.name} — install python-docx/openpyxl to extract text]"

    return f"[Unsupported format: {path.name}]"
